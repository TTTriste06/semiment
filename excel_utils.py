from datetime import datetime, timedelta
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill


def adjust_column_width(writer, sheet_name, df):
    """
    自动调整指定 sheet 的列宽（适用于 openpyxl）。
    """
    worksheet = writer.sheets[sheet_name]
    for idx, col in enumerate(df.columns, 1):
        max_len = df[col].astype(str).str.len().max()
        header_len = len(str(col))
        width = max(max_len, header_len) * 1.2 + 5
        worksheet.column_dimensions[get_column_letter(idx)].width = min(width, 50)


def auto_adjust_column_width_by_worksheet(ws):
    """
    自动根据 worksheet 的内容调整所有列宽（用于已写入的 worksheet）。
    """
    for idx, col in enumerate(ws.columns, 1):
        col_letter = get_column_letter(idx)
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    cell_len = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
                    max_length = max(max_length, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 5


def add_black_border(ws, row_count, col_count):
    """
    给前 row_count 行、col_count 列添加黑色边框（openpyxl worksheet）。
    """
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=row_count, min_col=1, max_col=col_count):
        for cell in row:
            cell.border = border


def process_date_column(df, date_col, date_format="%Y-%m"):
    """
    统一处理 Excel 中的日期列，支持日期序列/字符串 → datetime。
    添加新列 {date_col}_年月。
    """
    if pd.api.types.is_numeric_dtype(df[date_col]):
        df[date_col] = df[date_col].apply(lambda x: datetime(1899, 12, 30) + timedelta(days=float(x)))
    else:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    df[f'{date_col}_年月'] = df[date_col].dt.strftime(date_format)
    return df
