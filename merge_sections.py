import pandas as pd
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from excel_utils import adjust_column_width, add_black_border


def merge_safety_inventory(summary_df, df_safety, summary_sheet):
    """
    将安全库存数据合并到汇总 sheet，标红未匹配行。
    """
    df_safety.rename(columns={
        'WaferID': '晶圆品名',
        'OrderInformation': '规格',
        'ProductionNO.': '品名'
    }, inplace=True)
    df_safety['已匹配'] = False

    merged_df = summary_df.merge(
        df_safety[['晶圆品名', '规格', '品名', ' InvWaf', ' InvPart']],
        on=['晶圆品名', '规格', '品名'], how='left'
    )
    matched_index = df_safety.set_index(['晶圆品名', '规格', '品名']).index.isin(
        merged_df.dropna(subset=[' InvWaf', ' InvPart']).set_index(['晶圆品名', '规格', '品名']).index
    )
    df_safety.loc[matched_index, '已匹配'] = True

    # 写入合并列
    summary_sheet.merge_cells('D1:E1')
    summary_sheet['D1'] = '安全库存'
    summary_sheet['D1'].alignment = Alignment(horizontal='center', vertical='center')
    summary_sheet['D2'] = 'InvWaf（片）'
    summary_sheet['D2'].alignment = Alignment(horizontal='center', vertical='center')
    summary_sheet['E2'] = 'InvPart'
    summary_sheet['E2'].alignment = Alignment(horizontal='center', vertical='center')

    return merged_df, df_safety


def merge_unfulfilled_orders(summary_sheet, pending_df, start_col):
    """
    将未交订单数量（历史和未来）写入汇总 sheet。
    """
    pending_cols = [col for col in pending_df.columns if '未交订单数量_' in col]
    all_pending_cols = ['历史未交订单数量'] + pending_cols if '历史未交订单数量' in pending_df.columns else pending_cols
    pending_df['总未交订单'] = pending_df[all_pending_cols].sum(axis=1)

    columns_to_write = ['总未交订单'] + all_pending_cols

    end_col = start_col + len(columns_to_write) - 1
    summary_sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    summary_sheet.cell(row=1, column=start_col, value='未交订单').alignment = Alignment(horizontal='center', vertical='center')

    # 写入标题
    for idx, col in enumerate(columns_to_write, start=start_col):
        summary_sheet.cell(row=2, column=idx, value=col).alignment = Alignment(horizontal='center', vertical='center')

    return columns_to_write


def merge_prediction_data(summary_sheet, df_pred, summary_df):
    """
    将预测表中的数量与金额写入汇总表，标红未匹配行。
    """
    df_pred.columns = df_pred.iloc[0]
    df_pred = df_pred.drop([0, 0]).reset_index(drop=True)
    df_pred['已匹配'] = False

    required_cols = ['晶圆品名', '产品型号', 'ProductionNO.', '合计数量', '合计金额']
    if not all(col in df_pred.columns for col in required_cols):
        return df_pred

    start_col = summary_sheet.max_column + 1
    summary_sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col+1)
    summary_sheet.cell(row=1, column=start_col, value='预测').alignment = Alignment(horizontal='center', vertical='center')
    summary_sheet.cell(row=2, column=start_col, value='合计数量').alignment = Alignment(horizontal='center', vertical='center')
    summary_sheet.cell(row=2, column=start_col+1, value='合计金额').alignment = Alignment(horizontal='center', vertical='center')

    for row_idx in range(3, summary_sheet.max_row + 1):
        summary_key = [
            summary_sheet.cell(row=row_idx, column=1).value,
            summary_sheet.cell(row=row_idx, column=2).value,
            summary_sheet.cell(row=row_idx, column=3).value
        ]
        match = df_pred[
            (df_pred['晶圆品名'].astype(str) == str(summary_key[0])) &
            (df_pred['产品型号'].astype(str) == str(summary_key[1])) &
            (df_pred['ProductionNO.'].astype(str) == str(summary_key[2]))
        ]
        if not match.empty:
            qty, amt = match['合计数量'].values[0], match['合计金额'].values[0]
            summary_sheet.cell(row=row_idx, column=start_col, value=qty)
            summary_sheet.cell(row=row_idx, column=start_col+1, value=amt)
            df_pred.loc[match.index, '已匹配'] = True

    return df_pred


def mark_unmatched_rows(ws, df, start_row=2, color="FF0000"):
    """
    在 worksheet 中标红未匹配的 DataFrame 行。
    """
    red_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for row_idx, matched in enumerate(df['已匹配'], start=start_row):
        if not matched:
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = red_fill
